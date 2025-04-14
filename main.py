import requests
import hmac
import hashlib
import time
import urllib.parse
import random
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Binance API 信息
api_key = "自行填入"
secret_key = "自行填入"
base_url = "https://api.binance.com"
endpoint_path = '/sapi/v1/capital/withdraw/apply'

# 配置参数
COIN = "ETH"  # 提币币种
NETWORK = "OPTIMISM"  # 提币网络
AMOUNT_MIN = 0.01
AMOUNT_MAX = 0.02
DELAY_MIN = 60  # 最小延迟时间（秒）
DELAY_MAX = 600  # 最大延迟时间（秒）

# 记录提币结果
transaction_records = []

# 动态倒计时函数
def countdown(seconds):
    for remaining in range(seconds, 0, -1):
        sys.stdout.write(f"\r等待 {remaining} 秒后执行下一次提币...")
        sys.stdout.flush()
        time.sleep(1)
    sys.stdout.write("\r执行下一次提币！\n")
    sys.stdout.flush()

# 批量读取钱包地址
def read_wallets(file_path="wallets.txt"):
    try:
        with open(file_path, "r") as f:
            wallets = [line.strip() for line in f if line.strip()]
        return wallets
    except Exception as e:
        print(f"读取钱包地址时发生异常: {str(e)}")
        return []

# 获取币种当前价格
def get_price(symbol):
    try:
        # 直接返回 USDT 价格为 1.00
        if symbol.upper() == "USDT":
            return 1.00

        url = f"https://api.binance.com/api/v3/ticker/price?symbol={symbol}USDT"
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            return float(data['price'])
        else:
            print(f"获取价格失败: {response.text}")
            return None
    except Exception as e:
        print(f"获取价格时发生异常: {str(e)}")
        return None

# 提币请求
def withdraw(coin, network, address, amount):
    timestamp = round(time.time() * 1000)
    params = {
        "coin": coin,
        "network": network,
        "address": address,
        "amount": amount,
        "timestamp": timestamp
    }
    
    querystring = urllib.parse.urlencode(params)
    signature = hmac.new(secret_key.encode('utf-8'), msg=querystring.encode('utf-8'), digestmod=hashlib.sha256).hexdigest()
    url = base_url + endpoint_path + "?" + querystring + "&signature=" + signature

    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-MBX-APIKEY': api_key
    }

    try:
        response = requests.post(url, headers=headers)
        if response.status_code == 200:
            response_data = response.json()
            price = get_price(coin)
            usd_value = round(amount * price, 2) if price else "未知"
            print(f"成功提币 {amount} {coin} 到地址: {address}, ID: {response_data.get('id', '未知')}, 价值: {usd_value} USDT, 网络: {network}")
            transaction_records.append({
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "address": address,
                "coin": coin,
                "amount": amount,
                "value_usdt": usd_value,
                "network": network
            })
        else:
            print(f"提币失败 {amount} {coin} 到地址: {address} - 错误信息: {response.text}")
    except Exception as e:
        print(f"请求提币时发生异常: {str(e)}")

# 批量提币
def batch_withdraw(coin, network, wallets_file="wallets.txt"):
    wallets = read_wallets(wallets_file)
    if not wallets:
        print("没有找到钱包地址，退出程序。")
        return

    print(f"读取到 {len(wallets)} 个钱包地址，开始批量提币...")
    for idx, wallet in enumerate(wallets):
        amount = round(random.uniform(AMOUNT_MIN, AMOUNT_MAX), 8)  # 随机金额，精确到 8 位
        withdraw(coin, network, wallet, amount)
        
        # 如果不是最后一个钱包地址，执行倒计时
        if idx < len(wallets) - 1:
            delay = random.randint(DELAY_MIN, DELAY_MAX)  # 随机延迟时间
            countdown(delay)  # 动态倒计时

    # 保存记录到 Excel 文件
    save_to_excel(transaction_records, "withdrawal_records.xlsx")

# 保存记录到 Excel 文件
def save_to_excel(records, filename):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Withdrawal Records"
            headers = ["time", "address", "coin", "amount", "value_usdt", "network"]
            sheet.append(headers)

        # 写入数据
        for record in records:
            sheet.append([record["time"], record["address"], record["coin"], record["amount"], record["value_usdt"], record["network"]])

        # 自动调整列宽
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        # 保存文件
        workbook.save(filename)
        print(f"提币记录已追加到 {filename} 并自动调整列宽")
    except Exception as e:
        print(f"保存记录到文件时发生异常: {str(e)}")

# 主函数
if __name__ == "__main__":
    batch_withdraw(COIN, NETWORK, wallets_file="wallets.txt")
